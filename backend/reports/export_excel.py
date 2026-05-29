import io
from datetime import datetime

from database.models import query_detections, get_summary_stats, get_camera_analytics

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_excel(page=1, limit=10000, status=None, defect_type=None, camera_source=None,
                   start_date=None, end_date=None):
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed. Run: cd backend && pip install --user --break-system-packages openpyxl")

    records, _ = query_detections(
        page=1, limit=limit, status=status, defect_type=defect_type,
        camera_source=camera_source, start_date=start_date, end_date=end_date,
        sort_by="timestamp", sort_order="desc"
    )

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    ws = wb.active
    ws.title = "Detection History"

    headers = ["ID", "Timestamp", "Bottle ID", "Status", "Defect Type", "Confidence", "Camera Source"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for row_idx, r in enumerate(records, 2):
        values = [
            r["id"],
            r["timestamp"],
            r["bottle_id"],
            r["status"],
            r["defect_type"],
            f'{r["confidence"]:.2%}' if isinstance(r["confidence"], float) else r["confidence"],
            r["camera_source"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths = [8, 22, 14, 10, 20, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:G{len(records) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    stats = get_summary_stats()
    summary_data = [
        ("Metric", "Value"),
        ("Total Bottles", stats["total_bottles"]),
        ("Passed", stats["passed_bottles"]),
        ("Defective", stats["defective_bottles"]),
        ("Detection Accuracy", f'{stats["detection_accuracy"]}%'),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_b.font = header_font
            cell_b.fill = header_fill
        cell_a.border = thin_border
        cell_b.border = thin_border
        cell_a.alignment = Alignment(horizontal="center", vertical="center")
        cell_b.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
